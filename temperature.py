import requests
from bs4 import BeautifulSoup
import datetime
import openpyxl

def last_day_of_month(any_day):
    next_month = any_day.replace(day=28) + datetime.timedelta(days=4)
    return next_month - datetime.timedelta(days=next_month.day)

res = requests.get('https://pogoda1.ru/katalog/sverdlovsk-oblast/temperatura-vody/')
soup = BeautifulSoup(res.content, 'html.parser')
arr = []
for table in soup.select('.x-table > .x-row'):
	arr.append([])
	temperature = table.select_one('.x-cell-water-temp').get_text(strip=True)
	links = [a for a in table.select('.x-cell > .link')]
	name = links[0].text
	temperature = float(temperature)
	arr[-1].append(temperature)
	arr[-1].append(name)
midle_temperature = 0
for i in range(len(arr)):
	midle_temperature += arr[i][0]
midle_temperature = midle_temperature / len(arr)
month = (datetime.date.today().strftime('%B'))
day = (datetime.date.today().strftime('%d'))
number_month = int(datetime.date.today().strftime('%m'))
count_day_month = last_day_of_month(datetime.date(2020, number_month, 1))
midle_temperature_month = 0
f = open('Temperature.txt', 'a+')
f.write(day + ' of ' + month + ': ' + str(midle_temperature) + '\n')
if (count_day_month == day):
  count = 0
  f.seek(0)
  if i in range(len(f)):
    count += 1
    number = i.split()
    midle_temperature_month += float(number)
  midle_temperature_month = midle_temperature_month / count
  f.write(midle_temperature_month)
f.close()

wb = openpyxl.load_workbook('Temperature.xlsx')
sheet = wb['Лист1']
if sheet.cell(row = 1, column=1).value == None:
  sheet.cell(row = 1, column=1).value = "Date"
  sheet.cell(row = 1, column=2).value = "Average daily temperature in °С"
row_count = sheet.max_row
sheet.insert_rows(idx = row_count, amount = 1)
sheet.cell(row = row_count, column = 1).value = day + ' of ' + month
sheet.cell(row =row_count, column = 1).value = str(midle_temperature)
if (count_day_month == day):
  row_count = sheet.max_row
  sheet.insert_rows(idx = row_count, amount = 1)
  sheet.cell(row = row_count, column = 1).value = 'average monthly temperature: '
  sheet.cell(row = row_count, column = 2).value = midle_temperature_month
